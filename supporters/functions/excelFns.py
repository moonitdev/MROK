# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ##@@ brief:: generate & write excel file
# ##@@ note:: 
def write_excel(data=[], wb={}, callback=None):
    """
    기능: 
    입력:
        - 변수명 || 의미 | 데이터 타입 | 디폴트값 | 예시
        - 
        - 
    출력:
        - 의미 | 데이터 타입 | 예시
        - 
    Note:
        -
    """
    # create workbook open
    #load_wb = load_workbook()
    write_wb = Workbook()
    # worksheet open
    write_ws = write_wb.active
    write_ws.title = '성능평가'
    #write_ws = write_wb.create_sheet('충남')
    # write_ws = write_wb.active    # default sheet[Sheet1]

    # callback func: work with data
    if callback:
        callback(write_ws, data, wb)

        # write field name at cell
        # write field data at cell
    # save workbook
    write_wb.save('../_files/' + wb['name'])
    pass


def update_excel(data=[], wb={}, callback=None):
    """
    기능: 
    입력:
        - 변수명 || 의미 | 데이터 타입 | 디폴트값 | 예시
        - 
        - 
    출력:
        - 의미 | 데이터 타입 | 예시
        - 
    Note:
        -
    """
    # create workbook open
    #load_wb = load_workbook()
    write_wb = Workbook()
    # worksheet open
    
        # write field name at cell
        # write field data at cell
    # close workbook
    pass


# ##@@ brief:: generate excel file
# ##@@ note:: 
def read_excel(data=[], wb={}, callback=None):
    # create workbook open
    #load_wb = load_workbook()
    write_wb = Workbook()
    # worksheet open
    
        # write field name at cell
        # write field data at cell
    # close workbook
    pass